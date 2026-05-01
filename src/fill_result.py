from __future__ import annotations

from pathlib import Path
from copy import copy

import pandas as pd
from openpyxl import load_workbook


def fill_result_template(template_path: Path, output_path: Path, tasks: pd.DataFrame, max_rows: int = 9) -> None:
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    fill_rows = []
    for r in range(2, ws.max_row + 1):
        if isinstance(ws.cell(row=r, column=1).value, int):
            fill_rows.append(r)
    fill_rows = fill_rows[:max_rows]
    for row in fill_rows:
        for col in range(2, 6):
            ws.cell(row=row, column=col).value = None
    for i, (_, task) in enumerate(tasks.head(len(fill_rows)).iterrows()):
        row = fill_rows[i]
        ws.cell(row=row, column=1).value = int(i + 1)
        ws.cell(row=row, column=2).value = str(task["目标编号"])
        ws.cell(row=row, column=3).value = str(task["任务"])
        ws.cell(row=row, column=4).value = float(task["开始准备时刻(s)"])
        ws.cell(row=row, column=5).value = float(task["任务执行时刻(s)"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def fill_result_template_v2(template_path: Path, output_path: Path, events: pd.DataFrame) -> None:
    """Fill the result template without imposing the old visible 9-row capacity limit."""

    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    start_row = 2
    n = len(events)
    needed_last = start_row + n - 1
    if needed_last > ws.max_row:
        ws.insert_rows(ws.max_row + 1, amount=needed_last - ws.max_row)
    # Copy the original writable-row style downward.
    template_row = 2
    for r in range(start_row, needed_last + 1):
        for c in range(1, 6):
            src = ws.cell(template_row, c)
            dst = ws.cell(r, c)
            if src.has_style:
                dst._style = copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
    for r in range(start_row, needed_last + 1):
        for c in range(1, 6):
            ws.cell(r, c).value = None
    work = events.sort_values("execute_time").reset_index(drop=True)
    for i, (_, row) in enumerate(work.iterrows(), start=start_row):
        target_text = str(row.get("target_id", "")) if row.get("event_type") == "shooting" else str(row.get("covered_targets", ""))
        task_text = "射击" if row.get("event_type") == "shooting" else "拍照"
        ws.cell(i, 1).value = i - start_row + 1
        ws.cell(i, 2).value = target_text
        ws.cell(i, 3).value = task_text
        ws.cell(i, 4).value = float(row["start_time"])
        ws.cell(i, 5).value = float(row["execute_time"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
