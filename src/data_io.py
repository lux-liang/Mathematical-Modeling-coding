from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook


TIME_COL = "时间(s)"
X_COL = "X坐标(m)"
Y_COL = "Y坐标(m)"


@dataclass
class WorkbookInfo:
    file_name: str
    sheets: dict[str, pd.DataFrame]


def read_workbook(path: Path) -> WorkbookInfo:
    xl = pd.ExcelFile(path)
    sheets = {sheet: pd.read_excel(path, sheet_name=sheet) for sheet in xl.sheet_names}
    return WorkbookInfo(file_name=path.name, sheets=sheets)


def read_position_workbook(path: Path) -> dict[str, pd.DataFrame]:
    info = read_workbook(path)
    data: dict[str, pd.DataFrame] = {}
    for sheet, df in info.sheets.items():
        required = {TIME_COL, X_COL, Y_COL}
        missing = required.difference(df.columns)
        if missing:
            raise ValueError(f"{path.name}/{sheet} 缺少列: {sorted(missing)}")
        clean = df[[TIME_COL, X_COL, Y_COL]].copy()
        clean = clean.sort_values(TIME_COL).reset_index(drop=True)
        data[sheet] = clean
    return data


def read_targets(path: Path) -> pd.DataFrame:
    frames = []
    for sheet in pd.ExcelFile(path).sheet_names:
        df = pd.read_excel(path, sheet_name=sheet)
        if not {"编号", X_COL, Y_COL}.issubset(df.columns):
            raise ValueError(f"{path.name}/{sheet} 目标点列识别失败")
        task_type = "射击" if "射击" in sheet else "拍照"
        part = df[["编号", X_COL, Y_COL]].copy()
        part["任务"] = task_type
        frames.append(part)
    return pd.concat(frames, ignore_index=True)


def infer_result_slots(path: Path) -> tuple[str, list[int], list[str]]:
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
    rows = []
    for r in range(2, ws.max_row + 1):
        value = ws.cell(row=r, column=1).value
        if isinstance(value, int):
            rows.append(r)
    return ws.title, rows, [str(h) for h in headers]


def summarize_time(df: pd.DataFrame) -> dict[str, object]:
    t = df[TIME_COL].to_numpy(dtype=float)
    dt = np.diff(t)
    return {
        "start": float(np.nanmin(t)),
        "end": float(np.nanmax(t)),
        "count": int(len(t)),
        "median_dt": float(np.nanmedian(dt)) if len(dt) else np.nan,
        "min_dt": float(np.nanmin(dt)) if len(dt) else np.nan,
        "max_dt": float(np.nanmax(dt)) if len(dt) else np.nan,
        "duplicate_times": int(pd.Series(t).duplicated().sum()),
    }


def quality_summary(df: pd.DataFrame) -> dict[str, object]:
    numeric = df[[TIME_COL, X_COL, Y_COL]].apply(pd.to_numeric, errors="coerce")
    jumps = np.sqrt(np.diff(numeric[X_COL]) ** 2 + np.diff(numeric[Y_COL]) ** 2)
    if len(jumps) == 0:
        outlier_count = 0
        threshold = np.nan
    else:
        med = float(np.nanmedian(jumps))
        mad = float(np.nanmedian(np.abs(jumps - med)))
        threshold = med + 8.0 * 1.4826 * mad if mad > 1e-12 else med * 10.0
        outlier_count = int(np.sum(jumps > threshold))
    return {
        "missing_values": int(df.isna().sum().sum()),
        "duplicate_times": int(df[TIME_COL].duplicated().sum()) if TIME_COL in df else 0,
        "large_step_outliers": outlier_count,
        "large_step_threshold": threshold,
    }


def format_head(df: pd.DataFrame, n: int = 5) -> str:
    return "```text\n" + df.head(n).to_string(index=False) + "\n```"


def generate_data_report(base_dir: Path, output_path: Path, files: Iterable[str]) -> None:
    lines: list[str] = []
    lines.append("# 数据结构与质量分析报告")
    lines.append("")
    lines.append("说明：`Delta` 在后续模型中定义为将方式2时间轴修正为 `t2_aligned = t2 + Delta` 后与方式1对齐。")
    lines.append("")

    for file_name in files:
        path = base_dir / file_name
        info = read_workbook(path)
        lines.append(f"## {file_name}")
        lines.append(f"- Sheets: {', '.join(info.sheets.keys())}")
        if file_name == "result.xlsx":
            sheet, rows, headers = infer_result_slots(path)
            lines.append(f"- 模板工作表: {sheet}")
            lines.append(f"- 可填写区域: A-E 列，Excel 行号 {rows}，共 {len(rows)} 行")
            lines.append(f"- 左侧列名: {headers}")
        for sheet, df in info.sheets.items():
            lines.append(f"### {sheet}")
            lines.append(f"- 维度: {df.shape[0]} 行 x {df.shape[1]} 列")
            lines.append(f"- 列名: {list(df.columns)}")
            lines.append("")
            lines.append(format_head(df))
            lines.append("")
            if {TIME_COL, X_COL, Y_COL}.issubset(df.columns):
                ts = summarize_time(df)
                qs = quality_summary(df)
                lines.append(
                    f"- 时间范围: {ts['start']:.4f}s 到 {ts['end']:.4f}s；点数 {ts['count']}；"
                    f"采样间隔中位数 {ts['median_dt']:.4f}s，范围 [{ts['min_dt']:.4f}, {ts['max_dt']:.4f}]s"
                )
                lines.append(
                    f"- 质量: 缺失值 {qs['missing_values']}；重复时间 {qs['duplicate_times']}；"
                    f"明显大步长异常 {qs['large_step_outliers']} 个"
                )
            elif "编号" in df.columns:
                task_type = "射击" if "射击" in sheet else "拍照" if "拍照" in sheet else "未知"
                lines.append(f"- 目标点字段: 编号、任务类型={task_type}、坐标字段={X_COL}/{Y_COL}")
            lines.append("")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")
